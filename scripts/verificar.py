# -*- coding: utf-8 -*-
"""
Comprueba los entregables después de generarlos. No los lee "a ojo": recalcula
las sumas por su cuenta a partir de las facturas originales y las contrasta con
lo que salió en el Excel.

Uso:
    python scripts/verificar.py --mes Julio --anio 2026
    python scripts/verificar.py --mes Julio --anio 2026 --ejemplo

Devuelve 0 si todo está bien, 1 si hay algún fallo.
"""
import os
import re
import sys
import argparse

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")
    sys.stderr.reconfigure(errors="replace")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
import lector_facturas

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

SECCIONES = ["Resumen", "Diferencias de facturación", "Ejecución frente al plan",
             "Por cliente", "Dónde está la inversión", "Lo que no se pudo conciliar",
             "Cómo se ha medido"]

fallos = []


def comprobar(condicion, bien, mal):
    if condicion:
        print("  ✓ %s" % bien)
    else:
        print("  ✗ %s" % mal)
        fallos.append(mal)
    return condicion


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mes", required=True)
    p.add_argument("--anio", type=int, required=True)
    p.add_argument("--ejemplo", action="store_true")
    args = p.parse_args()

    mes_num = MESES.index(args.mes) + 1 if args.mes in MESES else 0
    sello = "%04d-%02d" % (args.anio, mes_num)
    ws = os.path.join(RAIZ, "workspace")
    ruta_x = os.path.join(ws, "%s-consolidado-pautas.xlsx" % sello)
    ruta_h = os.path.join(ws, "%s-informe-conciliacion.html" % sello)
    base = "ejemplos" if args.ejemplo else "entrada"

    print("Comprobando los entregables de %s %s" % (args.mes, args.anio))
    print()
    print("1. Los archivos existen y tienen contenido")
    for r in (ruta_x, ruta_h):
        existe = os.path.exists(r)
        tam = os.path.getsize(r) if existe else 0
        comprobar(existe and tam > 3000,
                  "%s (%s KB)" % (os.path.basename(r), tam // 1024),
                  "%s no existe o está casi vacío" % os.path.basename(r))
    if fallos:
        return terminar()

    print()
    print("2. El Excel tiene sus cuatro hojas")
    libro = openpyxl.load_workbook(ruta_x, data_only=True)
    for h in ("Consolidado", "Resumen por cliente", "Incidencias", "Cómo se ha medido"):
        comprobar(h in libro.sheetnames, "hoja '%s'" % h, "falta la hoja '%s'" % h)

    hoja = libro["Consolidado"]
    cab = [c.value for c in hoja[1]]
    filas = list(hoja.iter_rows(min_row=2, values_only=True))
    comprobar(len(filas) > 0, "el consolidado tiene %d campañas" % len(filas),
              "el consolidado está vacío")
    if not filas:
        return terminar()

    i_plan = cab.index("Presupuesto")
    i_cons = cab.index("Consumido")
    i_fact = cab.index("Facturado")
    i_ef = cab.index("Estado facturación")

    print()
    print("3. Las sumas, recalculadas aparte")
    suma_excel = sum(f[i_fact] or 0 for f in filas)

    lineas, _ = lector_facturas.leer_carpeta(os.path.join(RAIZ, base, "facturas"))
    reales = [l for l in lineas if not l.get("parcial")]
    cop = sum(l["importe"] for l in reales if l["divisa"] == "COP")
    otras = {}
    for l in reales:
        if l["divisa"] != "COP":
            otras[l["divisa"]] = otras.get(l["divisa"], 0.0) + l["importe"]

    # La TRM que se usó queda escrita en la hoja de metodología.
    trm_usada = None
    for fila in libro["Cómo se ha medido"].iter_rows(values_only=True):
        if fila and fila[0] == "Tipo de cambio" and fila[1]:
            m = re.match(r"([\d.]+)\s*COP/USD", str(fila[1]))
            if m:
                trm_usada = float(m.group(1))
    esperado = cop + sum(v * (trm_usada or 0) for v in otras.values())

    print("     Facturado en el Excel .......... %15.2f" % suma_excel)
    print("     Suma directa de las facturas ... %15.2f" % esperado)
    comprobar(abs(suma_excel - esperado) < 1.0,
              "el total facturado cuadra con la suma directa de las facturas",
              "el total facturado NO cuadra: Excel %.2f vs facturas %.2f"
              % (suma_excel, esperado))

    print()
    print("4. Los estados son coherentes con sus cifras")
    incoherentes = 0
    for f in filas:
        est, plan, cons, fact = f[i_ef], f[i_plan], f[i_cons], f[i_fact]
        if est == "CUADRA" and (not cons or fact is None
                                or abs((fact - cons) / cons) > 0.01):
            incoherentes += 1
        elif est == "SIN FACTURA" and fact is not None:
            incoherentes += 1
        elif est == "SIN PAUTA" and (plan is not None or cons is not None):
            incoherentes += 1
    comprobar(incoherentes == 0, "los %d estados cuadran con sus números" % len(filas),
              "%d campañas tienen un estado que no cuadra con sus cifras" % incoherentes)

    sin_explicar = sum(1 for f in filas if f[i_fact] is None
                       and f[i_ef] not in ("SIN FACTURA", "SIN DATOS"))
    comprobar(sin_explicar == 0, "ningún hueco se queda sin explicar",
              "%d campañas tienen el facturado vacío sin decir por qué" % sin_explicar)

    print()
    print("5. El informe")
    html = open(ruta_h, encoding="utf-8").read()
    for s in SECCIONES:
        marca = "<h2>%s</h2>" % s
        if marca not in html:
            comprobar(False, "", "falta la sección '%s'" % s)
            continue
        trozo = html.split(marca, 1)[1].split("<h2>", 1)[0]
        cuerpo = re.sub(r"<[^>]+>", "", trozo).strip()
        comprobar(len(cuerpo) > 40, "sección '%s' con contenido" % s,
                  "la sección '%s' está vacía" % s)

    # Marcadores de plantilla: solo en MAYÚSCULAS, para no confundir 'TODO' con
    # la palabra española 'todo'.
    for patron, desc, sensible in [(r"\[\[", "marcadores [[…]]", True),
                                   (r"\bTODO\b", "marcadores TODO", True),
                                   (r"\bLOREM\b", "texto de relleno", True),
                                   (r"<script", "JavaScript externo", False),
                                   (r"cdn\.", "enlaces a CDN", False),
                                   (r'<img[^>]+src="http', "imágenes externas", False)]:
        hits = re.findall(patron, html, 0 if sensible else re.I)
        comprobar(not hits, "sin %s" % desc,
                  "el informe tiene %s (%d veces)" % (desc, len(hits)))

    emojis = {c for c in html if ord(c) > 0x2190 and c not in "✓✗—…·"}
    comprobar(not emojis, "sin emojis",
              "el informe tiene emojis: %s" % " ".join(emojis))

    return terminar()


def terminar():
    print()
    if fallos:
        print("%d COMPROBACIONES FALLARON:" % len(fallos))
        for f in fallos:
            print("   ✗ %s" % f)
        print()
        print("Los entregables están en workspace/, pero revísalos antes de usarlos.")
        return 1
    print("TODAS LAS COMPROBACIONES PASARON.")
    print("Los entregables de workspace/ se pueden usar y enseñar.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
