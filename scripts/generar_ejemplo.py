# -*- coding: utf-8 -*-
"""
Genera el ejemplo de práctica: dos Excel de pauta y cinco facturas, todo inventado.

Los datos son de una agencia y unos clientes que no existen. Lleva errores metidos
a propósito, apuntados en ejemplos/LEEME.md, para comprobar que el kit los pilla.

Se ejecuta una sola vez al construir el kit:
    python scripts/generar_ejemplo.py
"""
import os
import openpyxl
from fpdf import FPDF

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EJ = os.path.join(RAIZ, "ejemplos")
PAUTAS = os.path.join(EJ, "pautas")
FACTURAS = os.path.join(EJ, "facturas")
for d in (PAUTAS, FACTURAS):
    if not os.path.isdir(d):
        os.makedirs(d)


# ---------------------------------------------------------------- Excel
def escribir_pauta(ruta, fila_cabecera, campanas, col_inicio=2):
    """Escribe un Excel de pauta imitando la plantilla real.

    fila_cabecera cambia entre archivos a propósito: en los reales estaba en la 7
    en un caso y en la 5 en otro.
    """
    libro = openpyxl.Workbook()
    libro.remove(libro.active)

    for mes, filas in campanas.items():
        hoja = libro.create_sheet(mes)
        hoja.cell(1, 2, "PLAN DE MEDIOS %s" % mes.upper())
        hoja.cell(2, 2, "Documento de trabajo interno")

        cabeceras = ["Campaña1", "Producto", "Campaña", "Fecha Inicio", "Fecha Final",
                     "Razón Social", "Medio", "Formato", "Estado", "Modelo de compra",
                     "Presupuesto Planeado", "Presupuesto Consumido",
                     "% cumplimiento Presupuesto", "Divisa"]
        for j, c in enumerate(cabeceras):
            hoja.cell(fila_cabecera, col_inicio + j, c)

        f = fila_cabecera + 1
        for fila in filas:
            for j, v in enumerate(fila):
                if v is not None:
                    hoja.cell(f, col_inicio + j, v)
            f += 1

        # Filas de total, como en los archivos reales. El kit tiene que ignorarlas.
        f += 1
        hoja.cell(f, col_inicio + 9, "SUBTOTAL")
        hoja.cell(f, col_inicio + 10,
                  sum(x[10] or 0 for x in filas if isinstance(x[10], (int, float))))
        hoja.cell(f + 1, col_inicio + 9, "IVA MEDIO")
        hoja.cell(f + 2, col_inicio + 9, "GRAN TOTAL")

    libro.save(ruta)
    return ruta


# Cliente 1 · VELARIA (cabecera en la fila 7)
velaria_julio = [
    # Campaña1, Producto, Campaña, Ini, Fin, RazónSocial, Medio, Formato, Estado, Modelo, Plan, Consumido, %, Divisa
    ["vel-master_2026-pos-search-cpc", "Máster", "Máster Marketing", "01/07/2026",
     "31/07/2026", "Google", "Search", "Texto", "Activo", "CPC", 4000000, 3980500,
     None, "COP"],
    ["vel-master_2026-pos-demandgen-cpc", "Máster", "Máster Marketing", "01/07/2026",
     "31/07/2026", "Google", "Demand Gen", "Vídeo", "Activo", "CPC", 2000000, 1750000,
     None, "COP"],
    ["vel-abierto_2026-pre-youtube-cpm", "Abiertos", "Programas Abiertos",
     "01/07/2026", "31/07/2026", "Google", "Youtube", "Preroll", "Activo", "CPM",
     1500000, 1499000, None, "COP"],
    # Sin factura: está en pauta y no aparece en ningún PDF.
    ["vel-abierto_2026-pre-meta-cpl", "Abiertos", "Programas Abiertos", "01/07/2026",
     "31/07/2026", "Meta", "Meta", "Multiformato", "Activo", "CPL", 900000, 880000,
     None, "COP"],
    # Hueco a propósito: sin presupuesto planeado.
    ["vel-becas_2026-pre-tiktok-cpc", "Becas", "Campaña Becas", "05/07/2026",
     "31/07/2026", "Tiktok", "TikTok", "Vídeo", "Activo", "CPC", None, 620000,
     None, "COP"],
    # Sobre-ejecución: gastó más de lo planeado.
    ["vel-exec_2026-pos-linkedin-cpl", "Executive", "Executive Education",
     "02/07/2026", "31/07/2026", "Linkedin", "LinkedIn", "Sponsored", "Activo", "CPL",
     1200000, 1455000, None, "COP"],
]

# Cliente 2 · NORTA (cabecera en la fila 5: estructura distinta a propósito)
norta_julio = [
    ["nor-grado_2026-pre-search-cpa", "Grado", "Grados", "01/07/2026", "31/07/2026",
     "Google", "Search", "Texto", "Activo", "CPA", 3000000, 2985000, None, "COP"],
    # Desviación de facturación grande: le cobran más de lo consumido.
    ["nor-grado_2026-pre-gdn-cpc", "Grado", "Grados", "01/07/2026", "31/07/2026",
     "Google", "GDN", "Display", "Activo", "CPC", 1000000, 940000, None, "COP"],
    # Hueco a propósito: sin consumido. No se puede juzgar.
    ["nor-posgrado_2026-pos-search-cpc", "Posgrado", "Posgrados", "01/07/2026",
     "31/07/2026", "Google", "Search", "Texto", "Pausada", "CPC", 800000, None,
     None, "COP"],
    ["nor-posgrado_2026-pos-tiktok-cpm", "Posgrado", "Posgrados", "01/07/2026",
     "31/07/2026", "Tiktok", "TikTok", "Vídeo", "Activo", "CPM", 700000, 681500,
     None, "COP"],
]

# Un mes anterior, para que el archivo tenga varias hojas como los reales.
velaria_junio = [
    ["vel-master_2026-pos-search-cpc", "Máster", "Máster Marketing", "01/06/2026",
     "30/06/2026", "Google", "Search", "Texto", "Activo", "CPC", 3500000, 3499000,
     None, "COP"],
]
norta_junio = [
    ["nor-grado_2026-pre-search-cpa", "Grado", "Grados", "01/06/2026", "30/06/2026",
     "Google", "Search", "Texto", "Activo", "CPA", 2800000, 2795000, None, "COP"],
]


# ---------------------------------------------------------------- PDFs
class Hoja(FPDF):
    def linea(self, txt, size=9, alto=5, negrita=False):
        self.set_font("Helvetica", "B" if negrita else "", size)
        self.cell(0, alto, txt, new_x="LMARGIN", new_y="NEXT")


def espaciar(texto):
    """Imita cómo Google Ads separa las letras en el PDF."""
    return " ".join(texto)


def factura_google(ruta):
    """Dos cuentas en un mismo PDF, y los nombres con las letras separadas."""
    p = Hoja()
    p.add_page()
    p.linea("Factura", 16, 9, True)
    p.linea("Numero de factura: 9900112233")
    p.linea("Fecha de la factura: 31 jul 2026")
    p.linea("Google LLC, 1600 Amphitheatre Pkwy, Mountain View, CA 94043")
    p.linea("Facturar a: Agencia Velaria SAS")
    p.linea("Importe total adeudado en COP  COP 11,224,500")
    p.ln(3)

    p.linea("ID de la cuenta: 111-222-3333")
    p.linea("Cuenta: Velaria Posgrados", negrita=True)
    p.linea("1 jul 2026 - 31 jul 2026")
    p.linea("Descripcion Cantidad Unidades Importe (COP)")
    for nombre, cant, uni, imp in [
            ("vel-master_2026-pos-search-cpc", "12500", "Clics", "3,980,000"),
            ("vel-master_2026-pos-demandgen-cpc", "980000", "Impresiones", "1,745,000"),
            ("vel-master_2026-pos-search-cpc", "40", "Clics", "500"),
            ("vel-abierto_2026-pre-youtube-cpm", "512000", "Impresiones", "1,499,000")]:
        p.linea("%s %s %s %s" % (espaciar(nombre), cant, uni, imp))
    p.linea("Actividad no valida: N. de la factura original: 9800111222, "
            "Mes de servicio original: jun 2026,")
    p.linea("Nombre de la campana: %s" % espaciar("vel-master_2026-pos-search-cpc"))
    p.linea("-1,200")
    p.linea("Subtotal en COP COP 7,223,300")
    p.ln(3)

    p.add_page()
    p.linea("ID de la cuenta: 444-555-6666")
    p.linea("Cuenta: Norta Universidad", negrita=True)
    p.linea("1 jul 2026 - 31 jul 2026")
    p.linea("Descripcion Cantidad Unidades Importe (COP)")
    for nombre, cant, uni, imp in [
            ("nor-grado_2026-pre-search-cpa", "9800", "Clics", "2,984,000"),
            # Desviacion plantada: consumido 940.000 pero facturan 1.010.000 (+7,4 %)
            ("nor-grado_2026-pre-gdn-cpc", "740000", "Impresiones", "1,010,000"),
            # Facturada sin estar en ninguna pauta
            ("nor-fantasma_2026-pre-search-cpc", "1200", "Clics", "450,000")]:
        p.linea("%s %s %s %s" % (espaciar(nombre), cant, uni, imp))
    p.linea("Subtotal en COP COP 4,444,000")
    p.output(ruta)
    return ruta


def factura_meta(ruta):
    p = Hoja()
    p.add_page()
    p.linea("Recibo para VELARIA_ABIERTOS (MERGE)", 12, 7, True)
    p.linea("Identificador de la cuenta: 3011223344556677")
    p.linea("Fecha de nota de pago pendiente/comprobante de pago")
    p.linea("31 jul. 2026, 9:14 p. m.")
    p.linea("Pagado")
    p.linea("Metodo de pago")
    p.linea("Visa .... 4417")
    p.linea("Tipo de producto")
    p.linea("Meta anuncios")
    p.linea("Campañas")
    # OJO: esta campana NO esta en la pauta de Velaria -> SIN PAUTA
    p.linea("vel-eventos_2026-pre-meta-cpl")
    p.linea("$ 315.400")
    p.linea("De 1 jul. 2026 a 31 jul. 2026")
    p.linea("Meta Platforms Ireland Limited")
    p.output(ruta)
    return ruta


def factura_tiktok(ruta):
    """Tabla de consumo con los nombres partidos en varias lineas, como la real."""
    p = Hoja()
    p.add_page()
    p.linea("TIKTOK INC.", 13, 7, True)
    p.linea("INVOICE")
    p.linea("Client Name Agencia Velaria SAS Invoice # MUUS20260099887")
    p.linea("Invoice Date August 01, 2026")
    p.linea("Billing Period : July 01, 2026 - July 31, 2026")
    p.ln(4)
    p.linea("Consumption Details:", 10, 6, True)
    p.ln(2)

    # Tabla dibujada con lineas para que extract_tables la reconozca.
    p.set_font("Helvetica", "B", 7)
    cols = [20, 28, 62, 18, 30, 30]
    cab = ["Statement", "Advertiser", "Campaign Name", "Target\nCountry",
           "Total Consumption\nin COP", "Cash Consumption\nCOP"]
    y0 = p.get_y()
    for w, t in zip(cols, cab):
        x0 = p.get_x()
        p.multi_cell(w, 5, t, border=1, new_x="RIGHT", new_y="TOP",
                     max_line_height=5)
        p.set_xy(x0 + w, y0)
    p.ln(11)

    p.set_font("Helvetica", "", 8)
    filas = [
        ("ST9911", "Velaria", "vel-becas_2026-pre-tiktok-cpc", "CO", "620,000.00",
         "620,000.00"),
        # Desviacion plantada: consumido 681.500 pero facturan 655.000 (-3,9 %)
        ("ST9911", "Norta", "nor-posgrado_2026-pos-tiktok-cpm", "CO", "655,000.00",
         "655,000.00"),
    ]
    for f in filas:
        y0 = p.get_y()
        for w, t in zip(cols, f):
            x0 = p.get_x()
            p.multi_cell(w, 6, t, border=1, new_x="RIGHT", new_y="TOP",
                         max_line_height=6)
            p.set_xy(x0 + w, y0)
        p.ln(12)
    p.output(ruta)
    return ruta


def factura_linkedin(ruta):
    """Factura en USD. El nombre unas veces va en la misma linea y otras debajo."""
    p = Hoja()
    p.add_page()
    p.linea("LinkedIn Ireland Unlimited Company INVOICE", 12, 7, True)
    p.linea("Invoice Number : 78990011223  Balance Due : USD 465.00")
    p.linea("Invoice Date : 02-AUG-2026")
    p.linea("PO Number or I/O Number : Velaria - Executive")
    p.linea("Currency : USD")
    p.linea("Billing Period : JUL2026")
    p.ln(3)
    p.linea("Invoice Details", 10, 6, True)
    p.linea("Line Description Qty Unit Price Billed Amount VAT Amount")
    # Caso A: nombre en la misma linea
    p.linea("1 Campaign: vel-exec_2026-pos-linkedin-cpl 465.0 1 465.00 0.00")
    p.linea("Sponsored Content : 1 of 1 0.00%")
    p.linea("Billing Period From 02-JUL-26 To 31-JUL-26")
    p.output(ruta)
    return ruta


def factura_tiktok_repetida(ruta):
    """Un segundo PDF de TikTok, DISTINTO del original (otro número de factura,
    otros bytes), que por error vuelve a facturar la misma campaña por el mismo
    importe exacto. No es el "duplicado exacto" (ese es una copia byte a byte
    del mismo archivo, ya probado abajo): esto es un archivo genuinamente
    distinto que coincide en campaña + importe, el caso que debe cazar el aviso
    de "posibles facturas duplicadas".
    """
    p = Hoja()
    p.add_page()
    p.linea("TIKTOK INC.", 13, 7, True)
    p.linea("INVOICE")
    p.linea("Client Name Agencia Velaria SAS Invoice # MUUS20260099888")
    p.linea("Invoice Date August 03, 2026")
    p.linea("Billing Period : July 01, 2026 - July 31, 2026")
    p.ln(4)
    p.linea("Consumption Details:", 10, 6, True)
    p.ln(2)
    p.set_font("Helvetica", "B", 7)
    cols = [20, 28, 62, 18, 30, 30]
    cab = ["Statement", "Advertiser", "Campaign Name", "Target\nCountry",
           "Total Consumption\nin COP", "Cash Consumption\nCOP"]
    y0 = p.get_y()
    for w, t in zip(cols, cab):
        x0 = p.get_x()
        p.multi_cell(w, 5, t, border=1, new_x="RIGHT", new_y="TOP",
                     max_line_height=5)
        p.set_xy(x0 + w, y0)
    p.ln(11)
    p.set_font("Helvetica", "", 8)
    # Mismo importe que la factura original de vel-becas: 620.000,00
    fila = ("ST9925", "Velaria", "vel-becas_2026-pre-tiktok-cpc", "CO",
            "620,000.00", "620,000.00")
    y0 = p.get_y()
    for w, t in zip(cols, fila):
        x0 = p.get_x()
        p.multi_cell(w, 6, t, border=1, new_x="RIGHT", new_y="TOP",
                     max_line_height=6)
        p.set_xy(x0 + w, y0)
    p.output(ruta)
    return ruta


def factura_desconocida(ruta):
    """Una plataforma que el kit NO conoce: tiene que decirlo, no inventarse nada."""
    p = Hoja()
    p.add_page()
    p.linea("PUBLIRED MEDIOS S.A.", 13, 7, True)
    p.linea("FACTURA DE VENTA No. PR-2026-4417")
    p.linea("Cliente: Agencia Velaria SAS")
    p.linea("Periodo: julio 2026")
    p.linea("Concepto: Pauta en vallas digitales    COP 2,400,000")
    p.linea("Total: COP 2,400,000")
    p.output(ruta)
    return ruta


# ---------------------------------------------------------------- principal
def main():
    print("Generando el ejemplo de práctica…")

    escribir_pauta(os.path.join(PAUTAS, "Nvo - VELARIA - 2026.xlsx"), 7,
                   {"Junio": velaria_junio, "Julio": velaria_julio})
    escribir_pauta(os.path.join(PAUTAS, "Nvo - NORTA - 2026.xlsx"), 5,
                   {"Junio": norta_junio, "Julio": norta_julio})
    print("   2 Excel de pauta")

    factura_google(os.path.join(FACTURAS, "9900112233.pdf"))
    factura_meta(os.path.join(FACTURAS,
                              "2026-07-31T21-14 Transaccion n 3011223344556677.pdf"))
    factura_tiktok(os.path.join(FACTURAS, "MUUS20260099887-Velaria-Invoice.pdf"))
    factura_tiktok_repetida(os.path.join(FACTURAS, "MUUS20260099888-Velaria-Invoice.pdf"))
    factura_linkedin(os.path.join(FACTURAS, "VELARIA LINKEDIN 78990011223.pdf"))
    factura_desconocida(os.path.join(FACTURAS, "PR-2026-4417 Publired.pdf"))

    # Duplicado exacto: el kit tiene que ignorarlo.
    origen = os.path.join(FACTURAS, "9900112233.pdf")
    with open(origen, "rb") as f:
        datos = f.read()
    with open(os.path.join(FACTURAS, "9900112233 (1).pdf"), "wb") as f:
        f.write(datos)
    print("   8 facturas (una es copia exacta de otra, y otra repite un importe "
          "en un archivo distinto, a propósito)")
    print("Listo. Están en ejemplos/pautas y ejemplos/facturas")


if __name__ == "__main__":
    main()
