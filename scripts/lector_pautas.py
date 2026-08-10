# -*- coding: utf-8 -*-
"""
Lee los Excel de pauta y devuelve una fila por campaña.

Las columnas se localizan por palabras clave dentro del encabezado, sin depender
de tildes, mayusculas ni de la posicion: vale cualquier cabecera que contenga la
palabra "campana", escrita como sea.

Regla de oro: NADA se adivina en silencio. Si una columna obligatoria no aparece,
esa hoja se descarta y se dice por que. Y cuando hay varias columnas de campana,
la tecnica se elige mirando los datos, no suponiendo.
"""
import os
import re
import glob
import unicodedata

import openpyxl

# Columnas que buscamos, por palabras clave dentro del nombre. La comparación se
# hace sin tildes y en minúsculas, así que da igual cómo esté escrito el encabezado:
# 'Campaña1', 'Campana', 'CAMPAÑA', 'Nombre de campaña' valen todos.
COLUMNAS = {
    "producto":   ["producto"],
    "plataforma": ["razonsocial", "razon social", "plataforma", "proveedor"],
    "medio":      ["medio"],
    "formato":    ["formato"],
    "estado":     ["estado"],
    "modelo":     ["modelodecompra", "modelo de compra"],
    "plan":       ["presupuesto planeado", "presupuestoplaneado", "planeado",
                   "presupuesto plan"],
    "consumido":  ["presupuesto consumido", "presupuestoconsumido", "consumido",
                   "ejecutado", "invertido"],
    "divisa":     ["divisa", "moneda"],
}

OBLIGATORIAS = ["campana", "plan", "consumido"]

# Lo que NO es una columna de campaña aunque lleve la palabra dentro.
NO_ES_CAMPANA = ("duracion", "estado de", "id de", "tipo de", "% ", "cumplimiento")

# Columnas que nunca deben capturarse: llevan las palabras clave pero son otra cosa
# ('% cumplimiento Presupuesto' no es el presupuesto).
NO_ES_IMPORTE = ("%", "cumplimiento", "costo por", "cpm", "cpc", "cpv", "cpl", "cps")

# Filas de resumen que viven dentro de la tabla y NO son campañas.
PALABRAS_TOTAL = ("subtotal", "iva medio", "iva", "gran total", "total", "totales")

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]


def sin_tildes(texto):
    """Quita tildes para comparar nombres de hoja sin sorpresas."""
    t = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in t if not unicodedata.combining(c))


def llave(nombre):
    """Llave de cruce: minúsculas y sin espacios.

    Los nombres de campaña de Google Ads no llevan espacios; el PDF los inserta
    al extraer el texto. Quitarlos en los dos lados hace que casen.
    """
    return re.sub(r"\s+", "", str(nombre)).lower()


def numero(valor):
    """Convierte a número. Devuelve None si no lo es: None significa SIN DATOS."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip()
    if not txt or txt in ("-", "--", "N/A", "n/a", "#N/A"):
        return None
    txt = txt.replace("$", "").replace("COP", "").replace("€", "").strip()
    # 1.234.567,89 -> 1234567.89   |   1,234,567.89 -> 1234567.89
    if "," in txt and "." in txt:
        if txt.rfind(",") > txt.rfind("."):
            txt = txt.replace(".", "").replace(",", ".")
        else:
            txt = txt.replace(",", "")
    elif "," in txt:
        txt = txt.replace(",", ".") if txt.count(",") == 1 else txt.replace(",", "")
    try:
        return float(txt)
    except ValueError:
        return None


def normalizar(texto):
    """Para comparar encabezados: sin tildes, en minúsculas, sin espacios de sobra."""
    return re.sub(r"\s+", " ", sin_tildes(texto)).strip().lower()


def parece_tecnico(valor):
    """¿Este valor parece un nombre técnico de campaña?

    Los nombres técnicos son etiquetas tipo 'vel-master_2026-pos-search-cpc':
    sin espacios y con guiones o guiones bajos. Los nombres legibles son
    'Máster en Marketing'. Esto permite elegir la columna correcta mirando los
    datos, y no fiándose de cómo se llame el encabezado.
    """
    if valor is None:
        return False
    t = str(valor).strip()
    if not t or " " in t:
        return False
    return ("-" in t or "_" in t) and len(t) >= 8


def columnas_de_campana(fila_cabecera, filas_datos):
    """Localiza la columna del nombre técnico de campaña, y la del legible.

    Acepta cualquier encabezado que contenga la palabra 'campaña' escrita como
    sea: Campaña1, Campana, CAMPAÑA, 'Nombre de campaña'…

    Si hay varias, decide mirando los datos: gana la que más valores tenga con
    pinta de nombre técnico. Devuelve (col_tecnica, col_legible).
    """
    candidatas = []
    for j, c in enumerate(fila_cabecera):
        if c is None:
            continue
        n = normalizar(c)
        if "campana" not in n:
            continue
        if any(x in n for x in NO_ES_CAMPANA):
            continue
        candidatas.append((j, n))

    if not candidatas:
        return None, None
    if len(candidatas) == 1:
        return candidatas[0][0], None

    # Varias candidatas: se cuenta cuántos valores de cada una parecen técnicos.
    puntos = {}
    for j, n in candidatas:
        aciertos = 0
        mirados = 0
        for fila in filas_datos:
            if j >= len(fila):
                continue
            v = fila[j]
            if v is None or not str(v).strip():
                continue
            mirados += 1
            if parece_tecnico(v):
                aciertos += 1
            if mirados >= 40:
                break
        puntos[j] = (aciertos / mirados) if mirados else 0.0

    orden = sorted(candidatas, key=lambda c: (-puntos[c[0]], c[0]))
    tecnica = orden[0][0]

    # Si ninguna tiene pinta técnica, se cae a la convención: la que acabe en
    # dígito ('Campaña1') o, si no, la primera.
    if puntos[tecnica] == 0:
        con_digito = [j for j, n in candidatas if re.search(r"\d\s*$", n)]
        tecnica = con_digito[0] if con_digito else candidatas[0][0]

    legible = next((j for j, n in candidatas if j != tecnica), None)
    return tecnica, legible


def buscar_cabecera(filas, limite=40):
    """Devuelve (indice_fila_1based, mapa_columnas) o (None, None).

    La cabecera es la primera fila que contiene una columna de campaña. Se busca
    así porque la fila cambia entre archivos (7 en un caso real, 5 en otro).
    """
    for i, fila in enumerate(filas[:limite], start=1):
        tecnica, legible = columnas_de_campana(fila, filas[i:i + 60])
        if tecnica is None:
            continue

        mapa = {"campana": tecnica}
        if legible is not None:
            mapa["campana_ver"] = legible

        for j, c in enumerate(fila):
            if c is None:
                continue
            n = normalizar(c)
            for clave, claves in COLUMNAS.items():
                if clave in mapa:
                    continue
                if clave in ("plan", "consumido") and any(x in n for x in NO_ES_IMPORTE):
                    continue
                if any(k in n for k in claves):
                    mapa[clave] = j
        return i, mapa
    return None, None


def hojas_de_mes(libro):
    """Nombres de hoja que son un mes, en orden de calendario."""
    encontradas = []
    for hoja in libro.sheetnames:
        limpio = sin_tildes(hoja).strip().lower()
        for mes in MESES:
            if limpio == sin_tildes(mes).lower():
                encontradas.append((MESES.index(mes), hoja))
    encontradas.sort()
    return [h for _, h in encontradas]


def leer_archivo(ruta, mes):
    """Lee una hoja de mes de un Excel de pauta.

    Devuelve (filas, incidencias). Si la hoja no se puede leer, filas va vacía y
    la incidencia dice exactamente por qué.
    """
    nombre_archivo = os.path.basename(ruta)
    incidencias = []
    try:
        libro = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    except Exception as e:
        return [], [{"archivo": nombre_archivo, "tipo": "No se pudo abrir",
                     "detalle": "%s: %s" % (type(e).__name__, e)}]

    # La hoja del mes pedido, comparando sin tildes ni mayúsculas.
    objetivo = None
    for hoja in libro.sheetnames:
        if sin_tildes(hoja).strip().lower() == sin_tildes(mes).strip().lower():
            objetivo = hoja
            break

    if objetivo is None:
        disponibles = hojas_de_mes(libro)
        libro.close()
        return [], [{"archivo": nombre_archivo, "tipo": "Sin hoja del mes",
                     "detalle": "No hay hoja '%s'. Meses en el archivo: %s"
                                % (mes, ", ".join(disponibles) or "ninguno")}]

    hoja = libro[objetivo]
    filas_crudas = list(hoja.iter_rows(values_only=True))
    libro.close()

    cab_i, mapa = buscar_cabecera(filas_crudas)
    if cab_i is None:
        return [], [{"archivo": nombre_archivo, "tipo": "Sin cabecera",
                     "detalle": "En las primeras 40 filas de la hoja '%s' no hay "
                                "ninguna columna cuyo nombre contenga la palabra "
                                "'campaña'. Se acepta escrita como sea: Campaña1, "
                                "Campana, CAMPAÑA, 'Nombre de campaña'…" % objetivo}]

    faltan = [c for c in OBLIGATORIAS if c not in mapa]
    if faltan:
        legibles = {"campana": "una columna de campaña",
                    "plan": "el presupuesto planeado",
                    "consumido": "el presupuesto consumido"}
        return [], [{"archivo": nombre_archivo, "tipo": "Faltan columnas",
                     "detalle": "En la hoja '%s' no se encontró %s"
                                % (objetivo, ", ni ".join(legibles[f] for f in faltan))}]

    def celda(fila, clave):
        j = mapa.get(clave)
        if j is None or j >= len(fila):
            return None
        return fila[j]

    salida = []
    for fila in filas_crudas[cab_i:]:
        bruto = celda(fila, "campana")
        if bruto is None or not str(bruto).strip():
            continue
        nombre = str(bruto).strip()
        if nombre.lower() in PALABRAS_TOTAL:
            continue

        salida.append({
            "llave": llave(nombre),
            "campana": nombre,
            "campana_ver": celda(fila, "campana_ver"),
            "producto": celda(fila, "producto"),
            "plataforma": celda(fila, "plataforma"),
            "medio": celda(fila, "medio"),
            "formato": celda(fila, "formato"),
            "estado": celda(fila, "estado"),
            "modelo": celda(fila, "modelo"),
            "divisa": celda(fila, "divisa"),
            "plan": numero(celda(fila, "plan")),
            "consumido": numero(celda(fila, "consumido")),
            "archivo": nombre_archivo,
            "hoja": objetivo,
            "cliente": cliente_desde_archivo(nombre_archivo),
        })

    return salida, incidencias


def cliente_desde_archivo(nombre_archivo):
    """El cliente sale del nombre del archivo de pauta.

    'Nvo - DECA - 2026.xlsx' -> 'DECA'
    'Nvo - S&O Admisiones Posgrados  Nacional - 2026.xlsx' -> 'S&O Admisiones Posgrados Nacional'
    """
    base = os.path.splitext(nombre_archivo)[0]
    base = re.sub(r"\(\d+\)\s*$", "", base).strip()      # copias '(1)'
    base = re.sub(r"^(Nvo|Nuevo)\s*-\s*", "", base, flags=re.I).strip()
    base = re.sub(r"\s*-\s*20\d\d\s*$", "", base).strip()  # año final
    base = re.sub(r"\s{2,}", " ", base)
    return base or os.path.splitext(nombre_archivo)[0]


def leer_carpeta(carpeta, mes):
    """Lee todos los Excel de pauta de una carpeta para un mes dado."""
    filas, incidencias = [], []
    patrones = ["*.xlsx", "*.xlsm"]
    rutas = []
    for p in patrones:
        rutas.extend(glob.glob(os.path.join(carpeta, p)))
    rutas = sorted(r for r in rutas if not os.path.basename(r).startswith("~$"))

    if not rutas:
        return [], [{"archivo": "(carpeta)", "tipo": "Carpeta vacía",
                     "detalle": "No hay ningún .xlsx en %s" % carpeta}]

    for ruta in rutas:
        f, inc = leer_archivo(ruta, mes)
        filas.extend(f)
        incidencias.extend(inc)
    return filas, incidencias
