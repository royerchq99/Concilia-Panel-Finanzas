# -*- coding: utf-8 -*-
"""
Cruza pauta y facturas, calcula el estado de cada campaña y escribe los dos
entregables en workspace/.

Uso:
    python scripts/conciliar.py --mes Julio --anio 2026
    python scripts/conciliar.py --mes Julio --anio 2026 --ejemplo
    python scripts/conciliar.py --mes Julio --anio 2026 --trm 3132.42

Reglas que no se rompen:
  - Lo que falta se marca SIN DATOS. Nunca se rellena con un cero ni con una
    estimación.
  - Si una campaña facturada no casa con ninguna de la pauta, va a incidencias
    para que decida una persona. El kit no adivina a cuál se parece.
  - Los importes salen de los archivos. Lo único que se calcula son las
    diferencias, y se dice cómo.
"""
import os
import sys
import argparse
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import lector_pautas
import lector_facturas
import trm as modulo_trm

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOLERANCIA = 0.01          # 1 %. Se puede cambiar aquí.
MESES = lector_pautas.MESES


PRODUCTO_POR_DEFECTO = "Conciliación de pautas"


def leer_marca():
    """Nombre del producto y firma del informe. Se configuran en marca.json."""
    ruta = os.path.join(RAIZ, "marca.json")
    producto, firma, pie = PRODUCTO_POR_DEFECTO, "", ""
    if os.path.exists(ruta):
        try:
            import json
            with open(ruta, encoding="utf-8") as f:
                d = json.load(f)
            producto = (d.get("producto") or "").strip() or PRODUCTO_POR_DEFECTO
            firma = (d.get("firma") or "").strip()
            pie = (d.get("pie") or "").strip()
        except Exception:
            pass
    return producto, firma, pie

# Dos preguntas distintas, dos estados distintos. Mezclarlos oculta lo importante:
# que una campaña gaste menos de lo planeado es normal y no dice nada de si la
# factura está bien.
ESTADOS_FACT = ["CUADRA", "DESVIACION EN FACTURACION",
                "SIN FACTURA", "SIN PAUTA", "SIN DATOS"]
ESTADOS_EJEC = ["EN PLAN", "POR DEBAJO DEL PLAN", "POR ENCIMA DEL PLAN", "SIN DATOS"]

COLORES = {
    "CUADRA": "C6EFCE",
    "DESVIACION EN FACTURACION": "FFC7CE",
    "SIN FACTURA": "DDEBF7",
    "SIN PAUTA": "FCE4D6",
    "SIN DATOS": "E7E6E6",
    "EN PLAN": "C6EFCE",
    "POR DEBAJO DEL PLAN": "FFEB9C",
    "POR ENCIMA DEL PLAN": "F8CBAD",
}


# ---------------------------------------------------------------- cálculo
def desviacion(valor, referencia):
    """Diferencia relativa. None si no se puede calcular."""
    if valor is None or referencia is None or referencia == 0:
        return None
    return (valor - referencia) / referencia


def estado_facturacion(plan, consumido, facturado):
    """¿Lo que se ejecutó es lo que se cobró? Es la pregunta crítica."""
    if facturado is None:
        if plan is None and consumido is None:
            return "SIN DATOS", "No hay ninguna cifra"
        return "SIN FACTURA", "Hay pauta pero no aparece en ninguna factura"

    if plan is None and consumido is None:
        return "SIN PAUTA", "Se facturó pero no está en ninguna hoja de pauta"

    d = desviacion(facturado, consumido)
    if d is None:
        return "SIN DATOS", "No hay consumido con el que comparar lo facturado"
    if abs(d) > TOLERANCIA:
        return ("DESVIACION EN FACTURACION",
                "Lo facturado se aparta un %+.1f %% de lo consumido" % (d * 100))
    return "CUADRA", "Facturado y consumido coinciden dentro del 1 %"


def estado_ejecucion(plan, consumido):
    """¿Se gastó lo que se dijo que se iba a gastar? Informativo, no es un error."""
    d = desviacion(consumido, plan)
    if d is None:
        return "SIN DATOS", "Falta el presupuesto planeado o el consumido"
    if abs(d) <= TOLERANCIA:
        return "EN PLAN", "Ejecutado según lo planeado (dentro del 1 %)"
    if d < 0:
        return ("POR DEBAJO DEL PLAN",
                "Se ejecutó un %.1f %% menos de lo planeado" % (abs(d) * 100))
    return ("POR ENCIMA DEL PLAN",
            "Se ejecutó un %.1f %% más de lo planeado" % (d * 100))


def conciliar(filas_pauta, lineas_factura, valor_trm, origen_trm):
    """Une pauta y facturas por la llave de campaña."""
    # Facturación agrupada por campaña (una campaña puede venir en varios PDFs).
    facturado, creditos, detalle = {}, {}, {}
    for l in lineas_factura:
        k = l["campana"]
        if l.get("parcial"):
            creditos[k] = creditos.get(k, 0.0) + l["credito"]
            continue
        importe, divisa = l["importe"], l["divisa"]
        convertido, nota = importe, ""
        if divisa != "COP":
            if valor_trm is None:
                convertido, nota = None, "sin TRM"
            else:
                convertido = importe * valor_trm
                nota = "%s %.2f x TRM %.2f" % (divisa, importe, valor_trm)
        facturado[k] = facturado.get(k, 0.0) + (convertido or 0.0)
        d = detalle.setdefault(k, {"plataformas": set(), "archivos": set(),
                                   "divisas": set(), "notas": set(), "recibos": 0,
                                   "lineas": []})
        d["plataformas"].add(l["plataforma"])
        d["archivos"].add(l["archivo"])
        d["divisas"].add(divisa)
        d["recibos"] += 1
        if nota:
            d["notas"].add(nota)
        if convertido is None:
            facturado[k] = None
        if importe is not None:
            d["lineas"].append((l["archivo"], divisa, round(importe, 2)))

    # Dos recibos de archivos distintos con el mismo importe exacto para la misma
    # campaña: no es el mismo caso que el "duplicado exacto" de lector_facturas.py
    # (esa es una copia byte a byte del mismo PDF). Aquí son archivos DISTINTOS que
    # casualmente casan en el importe — puede ser un cobro repetido, o dos cargos
    # legítimos que coinciden. No se descarta nada, solo se avisa para que alguien
    # lo mire.
    duplicados_posibles = {}
    for k, d in detalle.items():
        por_importe = {}
        for archivo, divisa, importe in d["lineas"]:
            por_importe.setdefault((divisa, importe), set()).add(archivo)
        archivos_en_duda = set()
        for (divisa, importe), archivos in por_importe.items():
            if len(archivos) > 1:
                archivos_en_duda |= archivos
        if archivos_en_duda:
            duplicados_posibles[k] = sorted(archivos_en_duda)

    resultado, incidencias = [], []
    llaves_pauta = set()

    for p in filas_pauta:
        k = p["llave"]
        llaves_pauta.add(k)
        fact = facturado.get(k)
        cred = creditos.get(k)
        d = detalle.get(k, {})
        est, motivo = estado_facturacion(p["plan"], p["consumido"], fact)
        est_e, motivo_e = estado_ejecucion(p["plan"], p["consumido"])
        resultado.append({
            "cliente": p["cliente"],
            "archivo": p["archivo"],
            "campana": p["campana"],
            "campana_ver": p["campana_ver"],
            "producto": p["producto"],
            "plataforma_pauta": p["plataforma"],
            "medio": p["medio"],
            "plataforma_factura": ", ".join(sorted(d.get("plataformas", []))) or None,
            "plan": p["plan"],
            "consumido": p["consumido"],
            "facturado": fact,
            "credito": cred,
            "divisa_factura": ", ".join(sorted(d.get("divisas", []))) or None,
            "conversion": "; ".join(sorted(d.get("notas", []))) or None,
            "recibos": d.get("recibos", 0),
            "desv_ejecucion": desviacion(p["consumido"], p["plan"]),
            "desv_facturacion": desviacion(fact, p["consumido"]),
            "estado": est,
            "motivo": motivo,
            "estado_ejec": est_e,
            "motivo_ejec": motivo_e,
            "posible_duplicado": k in duplicados_posibles,
            "archivos_duplicados": duplicados_posibles.get(k, []),
        })

    # Facturado que no casa con ninguna pauta.
    huerfanas, importe_huerfano = 0, 0.0
    plataformas_huerfanas = set()
    for k, importe in sorted(facturado.items()):
        if k in llaves_pauta:
            continue
        d = detalle.get(k, {})
        huerfanas += 1
        importe_huerfano += importe or 0.0
        plataformas_huerfanas |= d.get("plataformas", set())
        resultado.append({
            "cliente": None, "archivo": ", ".join(sorted(d.get("archivos", []))),
            "campana": k, "campana_ver": None, "producto": None,
            "plataforma_pauta": None, "medio": None,
            "plataforma_factura": ", ".join(sorted(d.get("plataformas", []))),
            "plan": None, "consumido": None, "facturado": importe,
            "credito": creditos.get(k),
            "divisa_factura": ", ".join(sorted(d.get("divisas", []))),
            "conversion": "; ".join(sorted(d.get("notas", []))) or None,
            "recibos": d.get("recibos", 0),
            "desv_ejecucion": None, "desv_facturacion": None,
            "estado": "SIN PAUTA",
            "motivo": "Se facturó pero no está en ninguna hoja de pauta del mes",
            "estado_ejec": "SIN DATOS",
            "motivo_ejec": "No hay pauta para esta campaña",
            "posible_duplicado": k in duplicados_posibles,
            "archivos_duplicados": duplicados_posibles.get(k, []),
        })

    # Una sola incidencia con el recuento, no una por campaña: con 81 líneas
    # sueltas la hoja de incidencias deja de leerse.
    if huerfanas:
        incidencias.append({
            "archivo": "(varias facturas)",
            "tipo": "Facturado sin pauta",
            "detalle": "%d campañas facturadas (%s) por un total de %s no aparecen en "
                       "ninguna pauta del mes. Normalmente significa que falta el Excel "
                       "de pauta de esos clientes en la carpeta. Están todas listadas "
                       "en la hoja Consolidado con estado SIN PAUTA."
                       % (huerfanas, ", ".join(sorted(plataformas_huerfanas)),
                          "{:,.0f}".format(importe_huerfano).replace(",", ".")),
        })
    return resultado, incidencias


# ---------------------------------------------------------------- Excel
def escribir_excel(ruta, filas, incidencias, contexto):
    libro = openpyxl.Workbook()

    negrita = Font(bold=True, color="FFFFFF")
    fondo = PatternFill("solid", fgColor="44546A")

    # --- Hoja 1: Consolidado
    hoja = libro.active
    hoja.title = "Consolidado"
    cabeceras = ["Cliente", "Archivo origen", "Campaña1", "Campaña", "Producto",
                 "Plataforma", "Medio", "Presupuesto", "Consumido", "Facturado",
                 "Crédito aplicado", "Divisa factura", "Conversión",
                 "Desv. ejecución %", "Estado ejecución",
                 "Desv. facturación %", "Estado facturación", "Motivo"]
    hoja.append(cabeceras)
    for c in range(1, len(cabeceras) + 1):
        hoja.cell(1, c).font = negrita
        hoja.cell(1, c).fill = fondo
        hoja.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)

    def pct(v):
        return None if v is None else round(v * 100, 2)

    for f in filas:
        hoja.append([
            f["cliente"], f["archivo"], f["campana"], f["campana_ver"], f["producto"],
            f["plataforma_factura"] or f["plataforma_pauta"], f["medio"],
            f["plan"], f["consumido"], f["facturado"], f["credito"],
            f["divisa_factura"], f["conversion"],
            pct(f["desv_ejecucion"]), f["estado_ejec"],
            pct(f["desv_facturacion"]), f["estado"], f["motivo"],
        ])

    col_ejec = cabeceras.index("Estado ejecución") + 1
    col_fact = cabeceras.index("Estado facturación") + 1
    for i in range(2, hoja.max_row + 1):
        for col in (col_ejec, col_fact):
            est = hoja.cell(i, col).value
            if est in COLORES:
                hoja.cell(i, col).fill = PatternFill("solid", fgColor=COLORES[est])
        for c in (8, 9, 10, 11):
            hoja.cell(i, c).number_format = "#,##0.00"
        for c in (14, 16):
            hoja.cell(i, c).number_format = "0.00"

    anchos = [22, 30, 46, 24, 16, 12, 14, 15, 15, 15, 15, 12, 26,
              15, 22, 16, 26, 46]
    for j, a in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(j)].width = a
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    # --- Hoja 2: Resumen por cliente
    resumen = libro.create_sheet("Resumen por cliente")
    resumen.append(["Cliente", "Campañas", "Presupuesto", "Consumido", "Facturado",
                    "% ejecución", "Cuadran", "Con desviación", "Sin factura"])
    for c in range(1, 10):
        resumen.cell(1, c).font = negrita
        resumen.cell(1, c).fill = fondo

    por_cliente = {}
    for f in filas:
        cl = f["cliente"] or "(sin pauta)"
        d = por_cliente.setdefault(cl, {"n": 0, "plan": 0.0, "cons": 0.0, "fact": 0.0,
                                        "ok": 0, "desv": 0, "sinf": 0})
        d["n"] += 1
        d["plan"] += f["plan"] or 0.0
        d["cons"] += f["consumido"] or 0.0
        d["fact"] += f["facturado"] or 0.0
        if f["estado"] == "CUADRA":
            d["ok"] += 1
        elif f["estado"].startswith("DESVIACION"):
            d["desv"] += 1
        elif f["estado"] == "SIN FACTURA":
            d["sinf"] += 1

    for cl in sorted(por_cliente):
        d = por_cliente[cl]
        eje = (d["cons"] / d["plan"] * 100) if d["plan"] else None
        resumen.append([cl, d["n"], d["plan"], d["cons"], d["fact"],
                        round(eje, 2) if eje is not None else None,
                        d["ok"], d["desv"], d["sinf"]])
    for i in range(2, resumen.max_row + 1):
        for c in (3, 4, 5):
            resumen.cell(i, c).number_format = "#,##0.00"
    for j, a in enumerate([28, 11, 16, 16, 16, 13, 10, 15, 12], start=1):
        resumen.column_dimensions[get_column_letter(j)].width = a
    resumen.freeze_panes = "A2"

    # --- Hoja 3: Incidencias
    inc = libro.create_sheet("Incidencias")
    inc.append(["Archivo", "Tipo", "Qué pasa"])
    for c in range(1, 4):
        inc.cell(1, c).font = negrita
        inc.cell(1, c).fill = fondo
    for i in incidencias:
        inc.append([i.get("archivo"), i.get("tipo"), i.get("detalle")])
    if not incidencias:
        inc.append(["", "Sin incidencias", "Todos los archivos se leyeron sin problemas"])
    for j, a in enumerate([34, 26, 110], start=1):
        inc.column_dimensions[get_column_letter(j)].width = a
    inc.freeze_panes = "A2"

    # --- Hoja 4: Cómo se ha medido
    met = libro.create_sheet("Cómo se ha medido")
    for clave, valor in contexto.items():
        met.append([clave, str(valor)])
    met.column_dimensions["A"].width = 34
    met.column_dimensions["B"].width = 96

    libro.save(ruta)
    return ruta


# ---------------------------------------------------------------- HTML
def dinero(v):
    if v is None:
        return '<span class="sindatos">SIN DATOS</span>'
    return "{:,.0f}".format(v).replace(",", ".")


def porcentaje(v):
    if v is None:
        return '<span class="sindatos">—</span>'
    return "%+.1f %%" % (v * 100)


def escribir_html(ruta, filas, incidencias, contexto, anterior=None):
    producto, firma, pie = leer_marca()
    total_plan = sum(f["plan"] or 0 for f in filas)
    total_cons = sum(f["consumido"] or 0 for f in filas)
    total_fact = sum(f["facturado"] or 0 for f in filas)

    cuenta = {}
    for f in filas:
        cuenta[f["estado"]] = cuenta.get(f["estado"], 0) + 1
    cuenta_ejec = {}
    for f in filas:
        cuenta_ejec[f["estado_ejec"]] = cuenta_ejec.get(f["estado_ejec"], 0) + 1

    # Cobertura de pauta: de todas las campañas vistas (con pauta o solo
    # facturadas), cuántas tienen su Excel de pauta cargado. No es una promesa
    # sobre "cuántos clientes deberían estar" —eso no está en ningún archivo—,
    # solo cuenta lo que sí se pudo ver.
    con_pauta = sum(1 for f in filas if f["cliente"] is not None)
    sin_pauta = cuenta.get("SIN PAUTA", 0)
    total_vistas = con_pauta + sin_pauta
    cobertura_pct = (con_pauta / total_vistas * 100) if total_vistas else 0

    duplicados = [f for f in filas if f.get("posible_duplicado")]

    n = len(filas)
    # El veredicto es sobre la facturación, y solo sobre lo que se PUEDE juzgar:
    # las campañas con pauta y con factura. Meter en el porcentaje las que no
    # tienen factura hundiría la nota sin que nada esté mal.
    comparables = cuenta.get("CUADRA", 0) + cuenta.get("DESVIACION EN FACTURACION", 0)
    conciliadas = cuenta.get("CUADRA", 0)
    pct_ok = (conciliadas / comparables * 100) if comparables else 0

    if not comparables:
        banda, texto_banda = "critica", "No se pudo conciliar nada"
        frase = ("Ninguna campaña tiene pauta y factura a la vez. Falta algún archivo: "
                 "revisa que estén los Excel de pauta y las facturas del mismo mes.")
    elif pct_ok >= 90:
        banda, texto_banda = "buena", "La facturación cuadra"
        frase = ("Casi todo lo que se ejecutó se cobró por el mismo importe. "
                 "No hay nada urgente que revisar en las facturas.")
    elif pct_ok >= 70:
        banda, texto_banda = "aceptable", "La facturación cuadra con excepciones"
        frase = ("La mayoría cuadra, pero hay campañas donde lo cobrado no coincide "
                 "con lo ejecutado. Míralas antes de aprobar el pago.")
    elif pct_ok >= 40:
        banda, texto_banda = "floja", "Hay bastante que revisar"
        frase = ("Más de una de cada tres campañas conciliables tiene diferencias "
                 "entre lo ejecutado y lo facturado.")
    else:
        banda, texto_banda = "critica", "La facturación no cuadra"
        frase = ("La mayor parte de lo facturado no coincide con lo ejecutado. "
                 "Antes de reclamar, comprueba que las facturas sean del mismo mes "
                 "que la pauta.")

    # Alertas ordenadas por dinero en juego.
    def dif_fact(f):
        if f["facturado"] is None or f["consumido"] is None:
            return 0.0
        return abs(f["facturado"] - f["consumido"])

    def dif_ejec(f):
        if f["consumido"] is None or f["plan"] is None:
            return 0.0
        return abs(f["consumido"] - f["plan"])

    alertas = sorted([f for f in filas
                      if f["estado"] == "DESVIACION EN FACTURACION"],
                     key=dif_fact, reverse=True)[:12]
    desvios_plan = sorted([f for f in filas
                           if f["estado_ejec"] in ("POR DEBAJO DEL PLAN",
                                                   "POR ENCIMA DEL PLAN")],
                          key=dif_ejec, reverse=True)[:12]

    # Reparto por plataforma, con cuántas campañas tiene cada una.
    plataformas, plataformas_n = {}, {}
    for f in filas:
        p = f["plataforma_factura"] or f["plataforma_pauta"] or "(sin plataforma)"
        plataformas[p] = plataformas.get(p, 0.0) + (f["consumido"] or 0.0)
        plataformas_n[p] = plataformas_n.get(p, 0) + 1
    tope = max(plataformas.values()) if plataformas else 1

    por_cliente = {}
    for f in filas:
        cl = f["cliente"] or "(sin pauta)"
        d = por_cliente.setdefault(cl, {"n": 0, "plan": 0.0, "cons": 0.0,
                                        "fact": 0.0, "ok": 0, "riesgo": 0.0})
        d["n"] += 1
        d["plan"] += f["plan"] or 0.0
        d["cons"] += f["consumido"] or 0.0
        d["fact"] += f["facturado"] or 0.0
        if f["estado"] == "CUADRA":
            d["ok"] += 1
        elif f["estado"] == "DESVIACION EN FACTURACION":
            d["riesgo"] += dif_fact(f)
        elif f["estado"] == "SIN FACTURA":
            d["riesgo"] += f["consumido"] or 0.0

    if anterior:
        for cl, d in por_cliente.items():
            fact_ant = anterior.get(cl)
            d["fact_ant"] = fact_ant
            d["cambio"] = desviacion(d["fact"], fact_ant) if fact_ant else None

    sin_medir = [f for f in filas
                 if f["estado"] in ("SIN FACTURA", "SIN PAUTA", "SIN DATOS")]

    h = []
    a = h.append
    a("""<meta charset="utf-8">
<title>%(producto)s · Conciliación de pautas · %(mes)s %(anio)s</title>
<style>
:root{--tinta:#1c2430;--suave:#5b6875;--linea:#dde3ea;--fondo:#fff;--panel:#f6f8fa;
--buena:#1e7a45;--aceptable:#8a6d1a;--floja:#a8541b;--critica:#a32020;}
*{box-sizing:border-box}
body{font-family:-apple-system,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
color:var(--tinta);background:var(--fondo);margin:0;padding:32px 20px;line-height:1.55}
.hoja{max-width:980px;margin:0 auto}
h1{font-size:26px;margin:0 0 4px}
h2{font-size:19px;margin:38px 0 12px;padding-bottom:7px;border-bottom:2px solid var(--linea)}
h3{font-size:15px;margin:22px 0 8px}
.sub{color:var(--suave);font-size:14px;margin:0 0 26px}
.veredicto{border:2px solid var(--linea);border-radius:10px;padding:22px 24px;margin:22px 0;background:var(--panel)}
.veredicto .cifra{font-size:44px;font-weight:700;line-height:1}
.veredicto .etiqueta{font-size:19px;font-weight:600;margin:6px 0 8px}
.veredicto .pie{margin-top:10px;font-size:13px;color:var(--suave)}
.buena .cifra,.buena .etiqueta{color:var(--buena)}
.aceptable .cifra,.aceptable .etiqueta{color:var(--aceptable)}
.floja .cifra,.floja .etiqueta{color:var(--floja)}
.critica .cifra,.critica .etiqueta{color:var(--critica)}
.rejilla{display:flex;flex-wrap:wrap;gap:12px;margin:18px 0}
.dato{flex:1 1 190px;border:1px solid var(--linea);border-radius:8px;padding:13px 15px;background:var(--panel)}
.dato .k{font-size:12px;color:var(--suave);text-transform:uppercase;letter-spacing:.4px}
.dato .v{font-size:21px;font-weight:600;margin-top:3px}
table{width:100%%;border-collapse:collapse;font-size:13.5px;margin:12px 0}
th,td{padding:7px 9px;border-bottom:1px solid var(--linea);text-align:left;vertical-align:top}
th{background:var(--panel);font-size:12px;text-transform:uppercase;letter-spacing:.4px;color:var(--suave)}
td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.campana{font-family:ui-monospace,Consolas,"Courier New",monospace;font-size:12px;word-break:break-all}
.tabla-scroll{overflow-x:auto}
.eti{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11.5px;font-weight:600;white-space:nowrap}
.e-cuadra{background:#d9f2e2;color:#155e33}
.e-ejec{background:#fdf0c9;color:#7a5f13}
.e-fact{background:#fbdcdc;color:#8f1d1d}
.e-sinf{background:#dfeaf7;color:#1f4f7a}
.e-sinp{background:#fbe3d4;color:#8a4418}
.e-sind{background:#e9e9e9;color:#555}
.sindatos{color:var(--suave);font-style:italic}
.barra{height:9px;background:#dfe5ec;border-radius:5px;overflow:hidden;margin-top:5px}
.barra span{display:block;height:100%%;background:#44546a}
.nota{background:var(--panel);border-left:3px solid var(--linea);padding:11px 15px;font-size:13.5px;margin:14px 0}
footer{margin-top:44px;padding-top:14px;border-top:1px solid var(--linea);
color:var(--suave);font-size:12.5px}
.barra-acciones{display:flex;justify-content:flex-end;gap:10px;margin-bottom:6px}
.boton{display:inline-flex;align-items:center;gap:7px;border:1px solid var(--linea);
background:var(--panel);color:var(--tinta);font:inherit;font-size:13.5px;
font-weight:600;padding:9px 16px;border-radius:7px;cursor:pointer;line-height:1}
.boton:hover{background:#eaeff5;border-color:#c3ccd6}
.boton:active{transform:translateY(1px)}
.pista{color:var(--suave);font-size:12px;text-align:right;margin:0 0 18px}
@media print{
  body{font-size:10.5pt;padding:0}
  .hoja{max-width:none}
  .noimprimir{display:none !important}
  h2{page-break-after:avoid}
  h3{page-break-after:avoid}
  table{page-break-inside:auto}
  tr{page-break-inside:avoid}
  thead{display:table-header-group}
  .seccion,.veredicto,.dato{page-break-inside:avoid}
  .tabla-scroll{overflow:visible}
  a[href^="http"]::after{content:" (" attr(href) ")"}
}
@page{margin:14mm 12mm}
</style>
<div class="hoja">
<div class="barra-acciones noimprimir">
<button class="boton" onclick="window.print()" type="button">
Descargar en PDF</button>
</div>
<p class="pista noimprimir">Se abre la ventana de impresión: elige
<strong>Guardar como PDF</strong> en el destino.</p>
<h1>%(producto)s</h1>
<p class="sub">Conciliación de pautas · %(mes)s %(anio)s ·
Generado el %(hoy)s%(firma_sub)s</p>
"""
      % dict(contexto, producto=producto,
               firma_sub=(" por " + firma) if firma else ""))

    a("""<div class="veredicto %s">
<div class="cifra">%.0f %%</div>
<div class="etiqueta">%s</div>
<div>%s</div>
<div class="pie">%d de %d campañas conciliables cuadran entre lo consumido y lo
facturado. Las %d que no tienen factura o no tienen pauta no cuentan en este
porcentaje: se listan aparte.</div>
</div>""" % (banda, pct_ok, texto_banda, frase, conciliadas, comparables,
             n - comparables))

    a('<div class="rejilla">')
    for k, v in (("Campañas", "{:,}".format(n).replace(",", ".")),
                 ("Presupuesto", dinero(total_plan)),
                 ("Consumido", dinero(total_cons)),
                 ("Facturado", dinero(total_fact))):
        a('<div class="dato"><div class="k">%s</div><div class="v">%s</div></div>' % (k, v))
    a("</div>")

    # Cobertura de pauta: cuántas de las campañas vistas este mes (con pauta o
    # solo facturadas) tienen su Excel de pauta cargado. No dice si faltan
    # clientes enteros —eso no está en ningún archivo—, solo lo que sí se ve.
    if total_vistas:
        a('<div class="dato" style="flex-basis:100%%;margin:4px 0 18px">'
          '<div class="k">Cobertura de pauta</div>'
          '<div class="v">%d de %d campañas (%.0f %%)</div>'
          '<div class="barra"><span style="width:%.1f%%"></span></div>'
          % (con_pauta, total_vistas, cobertura_pct, cobertura_pct))
        if sin_pauta:
            a('<div class="pie" style="text-align:left;margin-top:6px">%d campañas '
              'se facturaron sin tener pauta cargada — súbela y vuelve a conciliar '
              'para que se puedan comparar.</div>' % sin_pauta)
        a("</div>")

    # Resumen ejecutivo
    total_dif_ejec = total_cons - total_plan
    total_dif_fact = total_fact - total_cons
    a("<h2>Resumen</h2><p>")
    a("Se han cruzado <strong>%d campañas</strong> de <strong>%d clientes</strong>. " %
      (n, len(por_cliente)))
    a("Frente a un presupuesto de <strong>%s</strong> se ejecutaron <strong>%s</strong> " %
      (dinero(total_plan), dinero(total_cons)))
    a("(%s respecto al plan). " % porcentaje(desviacion(total_cons, total_plan)))
    a("Las facturas suman <strong>%s</strong>, %s sobre lo ejecutado. " %
      (dinero(total_fact), porcentaje(desviacion(total_fact, total_cons))))
    a("<strong>%d cuadran</strong> con su factura, %d tienen diferencias de "
      "facturación, %d no aparecen en ninguna factura y %d se facturaron sin estar "
      "en la pauta.</p>" %
      (cuenta.get("CUADRA", 0), cuenta.get("DESVIACION EN FACTURACION", 0),
       cuenta.get("SIN FACTURA", 0), cuenta.get("SIN PAUTA", 0)))
    a('<p class="nota">En dinero: la ejecución se separó del plan en <strong>%s</strong> '
      'y la facturación se separó de lo ejecutado en <strong>%s</strong>.</p>' %
      (dinero(abs(total_dif_ejec)), dinero(abs(total_dif_fact))))

    # Alertas de facturación: lo urgente
    a("<h2>Diferencias de facturación</h2>")
    a("<p>Campañas donde <strong>lo que se cobró no coincide con lo que se ejecutó</strong>. "
      "Es lo primero que hay que mirar, porque es dinero mal cobrado en un sentido o en "
      "el otro. Ordenado por dinero en juego, no por porcentaje.</p>")
    if not alertas:
        a("<p>Ninguna campaña se sale de la tolerancia del 1 %. "
          "Todo lo facturado coincide con lo ejecutado.</p>")
    else:
        a('<div class="tabla-scroll"><table><tr><th>Campaña</th><th>Cliente</th>'
          '<th class="num">Consumido</th><th class="num">Facturado</th>'
          '<th class="num">Diferencia</th><th class="num">%</th></tr>')
        for f in alertas:
            a('<tr><td class="campana">%s</td><td>%s</td>'
              '<td class="num">%s</td><td class="num">%s</td>'
              '<td class="num">%s</td><td class="num">%s</td></tr>' % (
                  f["campana"], f["cliente"] or "—",
                  dinero(f["consumido"]), dinero(f["facturado"]),
                  dinero(dif_fact(f)), porcentaje(f["desv_facturacion"])))
        a("</table></div>")

    # Posibles duplicados: dos recibos de archivos distintos con el mismo
    # importe exacto para la misma campaña. No se descarta nada —puede ser una
    # coincidencia real—, solo se avisa para que alguien lo mire antes de pagar.
    if duplicados:
        a("<h2>Posibles facturas duplicadas</h2>")
        a("<p>Estas campañas tienen dos o más recibos, de archivos distintos, "
          "por el <strong>mismo importe exacto</strong>. Puede ser un cobro "
          "repetido o dos cargos legítimos que coinciden — revísalo antes de "
          "aprobar el pago, no se ha descartado nada solo.</p>")
        a('<div class="tabla-scroll"><table><tr><th>Campaña</th><th>Cliente</th>'
          '<th class="num">Facturado</th><th>Archivos en duda</th></tr>')
        for f in duplicados:
            a('<tr><td class="campana">%s</td><td>%s</td>'
              '<td class="num">%s</td><td>%s</td></tr>' % (
                  f["campana"], f["cliente"] or "—", dinero(f["facturado"]),
                  ", ".join(f["archivos_duplicados"])))
        a("</table></div>")

    # Ejecución frente al plan: informativo
    a("<h2>Ejecución frente al plan</h2>")
    a("<p>Cuánto se gastó de lo que se había presupuestado. Gastar menos "
      "<strong>no es un error de facturación</strong>: es otra conversación, la de si "
      "la pauta se está ejecutando como se vendió.</p>")
    a('<div class="rejilla">')
    for etiqueta in ("EN PLAN", "POR DEBAJO DEL PLAN", "POR ENCIMA DEL PLAN"):
        a('<div class="dato"><div class="k">%s</div><div class="v">%d</div></div>'
          % (etiqueta.capitalize(), cuenta_ejec.get(etiqueta, 0)))
    a("</div>")
    if desvios_plan:
        a('<div class="tabla-scroll"><table><tr><th>Campaña</th><th>Cliente</th>'
          '<th class="num">Presupuesto</th><th class="num">Consumido</th>'
          '<th class="num">Diferencia</th><th class="num">%</th></tr>')
        for f in desvios_plan:
            a('<tr><td class="campana">%s</td><td>%s</td>'
              '<td class="num">%s</td><td class="num">%s</td>'
              '<td class="num">%s</td><td class="num">%s</td></tr>' % (
                  f["campana"], f["cliente"] or "—", dinero(f["plan"]),
                  dinero(f["consumido"]), dinero(dif_ejec(f)),
                  porcentaje(f["desv_ejecucion"])))
        a("</table></div>")

    # Por cliente — ordenado por dinero en juego (desviación + sin factura),
    # no por tamaño del cliente. A quién perseguir primero, no el más grande.
    a("<h2>Por cliente</h2>")
    a("<p>Ordenado por <strong>dinero en juego</strong>: la suma de lo que tiene "
      "diferencia de facturación más lo que se gastó y no aparece en ninguna "
      "factura. El cliente más arriba es el que conviene revisar primero.</p>")
    cabeceras_cl = ('<th>Cliente</th><th class="num">Campañas</th>'
                    '<th class="num">Presupuesto</th><th class="num">Consumido</th>'
                    '<th class="num">Facturado</th><th class="num">Ejecución</th>'
                    '<th class="num">Cuadran</th><th class="num">Dinero en juego</th>')
    if anterior:
        cabeceras_cl += '<th class="num">Vs. mes anterior</th>'
    a('<div class="tabla-scroll"><table><tr>%s</tr>' % cabeceras_cl)
    for cl in sorted(por_cliente, key=lambda c: -por_cliente[c]["riesgo"]):
        d = por_cliente[cl]
        eje = desviacion(d["cons"], d["plan"]) if d["plan"] else None
        fila = ('<tr><td>%s</td><td class="num">%d</td><td class="num">%s</td>'
                '<td class="num">%s</td><td class="num">%s</td><td class="num">%s</td>'
                '<td class="num">%d de %d</td><td class="num">%s</td>' % (
                    cl, d["n"], dinero(d["plan"]), dinero(d["cons"]), dinero(d["fact"]),
                    porcentaje(eje), d["ok"], d["n"],
                    dinero(d["riesgo"]) if d["riesgo"] else "—"))
        if anterior:
            fila += '<td class="num">%s</td>' % porcentaje(d.get("cambio"))
        a(fila + "</tr>")
    a("</table></div>")

    # Plataformas
    a("<h2>Dónde está la inversión</h2>")
    for p in sorted(plataformas, key=lambda k: -plataformas[k]):
        v = plataformas[p]
        ancho = (v / tope * 100) if tope else 0
        parte = (v / total_cons * 100) if total_cons else 0
        a('<h3>%s — %s (%.1f %%, %d campañas)</h3>'
          '<div class="barra"><span style="width:%.1f%%"></span></div>'
          % (p, dinero(v), parte, plataformas_n.get(p, 0), ancho))

    # Sin medir
    a("<h2>Lo que no se pudo conciliar</h2>")
    if not sin_medir:
        a("<p>Nada: todas las campañas tienen sus tres cifras.</p>")
    else:
        a("<p>Estas campañas no tienen las tres cifras. <strong>No se les ha puesto un "
          "cero ni una estimación</strong>: se listan para que decidas tú.</p>")
        a('<div class="tabla-scroll"><table><tr><th>Campaña</th><th>Cliente</th>'
          '<th>Estado</th><th>Por qué</th></tr>')
        for f in sin_medir[:60]:
            clase = {"SIN FACTURA": "e-sinf", "SIN PAUTA": "e-sinp",
                     "SIN DATOS": "e-sind"}[f["estado"]]
            a('<tr><td class="campana">%s</td><td>%s</td>'
              '<td><span class="eti %s">%s</span></td><td>%s</td></tr>' % (
                  f["campana"], f["cliente"] or "—", clase, f["estado"], f["motivo"]))
        a("</table></div>")
        if len(sin_medir) > 60:
            a("<p>… y %d más. Están todas en la hoja <em>Consolidado</em> del Excel.</p>"
              % (len(sin_medir) - 60))

    # Incidencias
    if incidencias:
        a("<h2>Incidencias al leer los archivos</h2>")
        a('<div class="tabla-scroll"><table><tr><th>Archivo</th><th>Tipo</th>'
          '<th>Qué pasa</th></tr>')
        for i in incidencias[:40]:
            a("<tr><td>%s</td><td>%s</td><td>%s</td></tr>" % (
                i.get("archivo", ""), i.get("tipo", ""), i.get("detalle", "")))
        a("</table></div>")

    # Cómo se ha medido
    a("<h2>Cómo se ha medido</h2><table>")
    for k, v in contexto.items():
        a("<tr><td><strong>%s</strong></td><td>%s</td></tr>" % (k, v))
    a("</table>")

    a("<footer>%s%s · Conciliación de pautas · %s %s<br>"
      "Los importes salen de los archivos de pauta y de las facturas. "
      "Lo único calculado son las diferencias y los porcentajes. "
      "Lo que falta se marca SIN DATOS.%s</footer></div>"
      % ((firma + " · ") if firma else "", producto,
         contexto["mes"], contexto["anio"],
         ("<br>" + pie) if pie else ""))

    with open(ruta, "w", encoding="utf-8") as f:
        f.write("\n".join(h))
    return ruta


# ------------------------------------------------------------ mes anterior
def leer_facturado_anterior(ruta):
    """Facturado por cliente del consolidado de un mes anterior, si existe.

    Solo lee ese archivo, nunca lo toca. Si no existe o no tiene la forma
    esperada, se devuelve None: sin comparación, no una comparación inventada.
    """
    if not os.path.exists(ruta):
        return None
    try:
        libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        if "Resumen por cliente" not in libro.sheetnames:
            return None
        hoja = libro["Resumen por cliente"]
        datos = {}
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if not fila or fila[0] is None:
                continue
            cliente, facturado = fila[0], fila[4] if len(fila) > 4 else None
            datos[cliente] = facturado
        return datos
    except Exception:
        return None


# ---------------------------------------------------------------- principal
def main():
    p = argparse.ArgumentParser(description="Concilia pauta y facturas de un mes")
    p.add_argument("--mes", required=True, help="Nombre de la hoja: Julio, Agosto…")
    p.add_argument("--anio", type=int, required=True)
    p.add_argument("--ejemplo", action="store_true",
                   help="Usa ejemplos/ en vez de entrada/")
    p.add_argument("--trm", type=float, default=None,
                   help="Tipo de cambio COP/USD, si no se quiere consultar")
    args = p.parse_args()

    base = "ejemplos" if args.ejemplo else "entrada"
    dir_pautas = os.path.join(RAIZ, base, "pautas")
    dir_facturas = os.path.join(RAIZ, base, "facturas")

    for d in (dir_pautas, dir_facturas):
        if not os.path.isdir(d):
            print("No existe la carpeta %s" % d)
            return 2

    print("Leyendo pautas de %s (hoja '%s')…" % (dir_pautas, args.mes))
    filas_pauta, inc_p = lector_pautas.leer_carpeta(dir_pautas, args.mes)
    print("   %d campañas de pauta, %d incidencias" % (len(filas_pauta), len(inc_p)))

    print("Leyendo facturas de %s…" % dir_facturas)
    lineas_fact, inc_f = lector_facturas.leer_carpeta(dir_facturas)
    reales = [l for l in lineas_fact if not l.get("parcial")]
    print("   %d líneas de factura, %d incidencias" % (len(reales), len(inc_f)))

    if not filas_pauta and not reales:
        print("\nNo hay nada que conciliar: ni pautas ni facturas legibles.")
        for i in inc_p + inc_f:
            print("   [%s] %s: %s" % (i["tipo"], i["archivo"], i["detalle"]))
        return 1

    # TRM: solo si hay algo en otra divisa.
    otras = sorted({l["divisa"] for l in reales if l["divisa"] != "COP"})
    valor_trm, origen_trm = None, "No hizo falta: todo estaba en COP"
    if otras:
        if args.trm is not None:
            valor_trm, origen_trm = args.trm, "Aportada a mano: %.2f" % args.trm
        else:
            mes_num = MESES.index(args.mes) + 1 if args.mes in MESES else None
            if mes_num:
                valor_trm, origen_trm = modulo_trm.trm_de_cierre(args.anio, mes_num)
            else:
                origen_trm = "No se reconoció el mes '%s'" % args.mes
        if valor_trm is None:
            print("\n   AVISO: hay facturas en %s y no se pudo obtener la TRM."
                  % ", ".join(otras))
            print("   Motivo: %s" % origen_trm)
            print("   Esas campañas quedarán como SIN DATOS. Para convertirlas, "
                  "vuelve a lanzarlo con --trm VALOR")
        else:
            print("   TRM aplicada: %.2f (%s)" % (valor_trm, origen_trm))

    filas, inc_c = conciliar(filas_pauta, lineas_fact, valor_trm, origen_trm)
    incidencias = inc_p + inc_f + inc_c

    contexto = {
        "mes": args.mes, "anio": args.anio,
        "hoy": datetime.date.today().strftime("%d/%m/%Y"),
        "Mes conciliado": "%s %s" % (args.mes, args.anio),
        "Origen de los datos": "%s/pautas y %s/facturas" % (base, base),
        "Campañas de pauta leídas": len(filas_pauta),
        "Líneas de factura leídas": len(reales),
        "Tolerancia": "1 % — por debajo de esa diferencia se considera que cuadra",
        "Tipo de cambio": origen_trm if valor_trm is None
                          else "%.2f COP/USD — %s" % (valor_trm, origen_trm),
        "Créditos": "Los créditos por actividad no válida van en columna aparte, "
                    "no restados del facturado",
        "Regla de los huecos": "Lo que falta se marca SIN DATOS. No se estima nunca",
    }

    salida = os.path.join(RAIZ, "workspace")
    if not os.path.isdir(salida):
        os.makedirs(salida)
    mes_num = MESES.index(args.mes) + 1 if args.mes in MESES else 0
    sello = "%04d-%02d" % (args.anio, mes_num)

    anterior = None
    if mes_num:
        mes_ant, anio_ant = (mes_num - 1, args.anio) if mes_num > 1 else (12, args.anio - 1)
        sello_ant = "%04d-%02d" % (anio_ant, mes_ant)
        anterior = leer_facturado_anterior(
            os.path.join(salida, "%s-consolidado-pautas.xlsx" % sello_ant))
        contexto["Comparación con el mes anterior"] = (
            "%s-consolidado-pautas.xlsx (columna Facturado de 'Resumen por cliente')"
            % sello_ant if anterior else
            "No hay consolidado de %s en workspace/, no se compara" % sello_ant)

    ruta_x = escribir_excel(os.path.join(salida, "%s-consolidado-pautas.xlsx" % sello),
                            filas, incidencias, contexto)
    ruta_h = escribir_html(os.path.join(salida, "%s-informe-conciliacion.html" % sello),
                           filas, incidencias, contexto, anterior)

    cuenta, cuenta_e = {}, {}
    for f in filas:
        cuenta[f["estado"]] = cuenta.get(f["estado"], 0) + 1
        cuenta_e[f["estado_ejec"]] = cuenta_e.get(f["estado_ejec"], 0) + 1

    print("\nLISTO")
    print("   Excel:   %s" % ruta_x)
    print("   Informe: %s" % ruta_h)
    print("\n   %d campañas. Facturación:" % len(filas))
    for e in ESTADOS_FACT:
        if cuenta.get(e):
            print("      %-30s %d" % (e, cuenta[e]))
    print("   Ejecución frente al plan:")
    for e in ESTADOS_EJEC:
        if cuenta_e.get(e):
            print("      %-30s %d" % (e, cuenta_e[e]))
    if incidencias:
        print("\n   %d incidencias (hoja 'Incidencias' del Excel)" % len(incidencias))
    return 0


if __name__ == "__main__":
    sys.exit(main())
